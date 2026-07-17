from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Finding, PersonResult


NAVY = "173B63"
BLUE = "1677FF"
PALE_BLUE = "EAF3FF"
TEXT = "243447"
MUTED = "64748B"
LINE = "DCE5EF"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(name="Microsoft YaHei UI", size=10, color="FFFFFF", bold=True)
TITLE_FONT = Font(name="Microsoft YaHei UI", size=16, color="FFFFFF", bold=True)
BODY_FONT = Font(name="Microsoft YaHei UI", size=10, color=TEXT)
THIN_BOTTOM = Border(bottom=Side(style="thin", color=LINE))

STATUS_FILL = {
    "通过": "C6EFCE", "比对通过": "C6EFCE", "一致": "C6EFCE", "齐全": "C6EFCE",
    "资料完整": "C6EFCE", "已填写": "C6EFCE", "已采用": "C6EFCE", "已计算": "C6EFCE",
    "在籍": "DDEBF7", "毕业": "DDEBF7", "豁免": "DDEBF7", "不适用": "E7E6E6",
    "待复核": "FFEB9C", "待人工复核": "FFEB9C", "人工复核": "FFEB9C",
    "无法核对": "FFEB9C", "存在缺填/未识别项": "FFEB9C",
    "退回": "FFC7CE", "不一致": "FFC7CE", "不通过": "FFC7CE",
    "资料不完整": "FFC7CE", "发现异常": "FFC7CE", "存在异常": "FFC7CE",
    "缺少材料": "FFC7CE", "缺少或不一致": "FFC7CE", "缺少信息": "FFEB9C",
    "材料不采用": "E7E6E6",
}

ERROR_STATUSES = {"退回", "不一致", "不通过", "缺少材料", "缺少或不一致"}
MISSING_STATUSES = {"缺少信息"}
REVIEW_STATUSES = {"待复核", "人工复核", "无法核对"}
PASS_STATUSES = {
    "通过", "一致", "齐全", "豁免", "已填写", "已采用", "已计算", "已检测",
    "在籍", "毕业", "不适用",
}
IGNORED_CONTENT_CATEGORIES = {"材料完整性", "材料质量", "材料择优", "人工预设"}

WIDTHS = {
    "人员": 13, "综合结论": 13, "完整性结论": 14, "内容核对结论": 18,
    "缺少/无合格材料": 28, "内容异常或复核事项": 48, "完整性说明": 50,
    "核对项目": 15, "核对事实": 52, "异常/复核事项": 52,
    "材料类型": 22, "状态": 14, "判断依据": 48, "数量或结果": 26,
    "类别": 19, "字段": 22, "说明": 52, "字段值": 52,
    "来源": 42, "来源文件与页码": 42, "来源文件": 32, "文件": 34,
    "企业名称": 34, "从事职业": 22, "企业名称说明": 42, "经营范围": 48,
    "质量原因": 42, "处理异常": 42, "退回/异常原因": 52,
}


def _sheet(wb: Workbook, title: str, headers: list[str], description: str):
    ws = wb.create_sheet(title)
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = description
    ws["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws["A2"].font = Font(name="Microsoft YaHei UI", size=9, color="285A8A")
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    ws.append([])
    ws.append(headers)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    for cell in ws[4]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28
    ws.auto_filter.ref = f"A4:{last_col}4"
    return ws


def _unique(values: list[str]) -> list[str]:
    result = []
    for value in values:
        if value and value not in result:
            result.append(value)
    return result


def _short_finding(finding: Finding) -> str:
    value = f"（{finding.values}）" if finding.values else ""
    return f"{finding.field}：{finding.status}{value}"


def _material_summary(result: PersonResult) -> dict[str, str | int]:
    findings = [finding for finding in result.findings if finding.category == "材料完整性"]
    missing = _unique([
        finding.field for finding in findings if finding.status in ERROR_STATUSES
    ])
    reviews = _unique([
        finding.field for finding in findings if finding.status in REVIEW_STATUSES | MISSING_STATUSES
    ])
    unused = _unique([
        material.path.name for material in result.materials
        if not material.selected_as_basis
        and (material.quality_status != "合格" or material.errors)
    ])
    selected_count = sum(material.selected_as_basis for material in result.materials)
    if missing:
        conclusion = "资料不完整"
        explanation = "缺少或未找到合格材料：" + "、".join(missing)
    else:
        conclusion = "资料完整"
        explanation = "所有已触发的必需材料均至少有1份合格可用"
    if reviews:
        explanation += "；需复核：" + "、".join(reviews)
    if unused:
        explanation += "；另有未采用副本：" + "、".join(unused[:6])
        if len(unused) > 6:
            explanation += f"等{len(unused)}份"
    return {
        "结论": conclusion,
        "缺少": "、".join(missing),
        "说明": explanation,
        "采用数": selected_count,
        "总数": len(result.materials),
    }


def _content_findings(result: PersonResult) -> list[Finding]:
    return [
        finding for finding in result.findings
        if finding.category not in IGNORED_CONTENT_CATEGORIES
        and finding.status != "材料不采用"
    ]


def _content_summary(result: PersonResult) -> dict[str, str]:
    findings = _content_findings(result)
    errors = [finding for finding in findings if finding.status in ERROR_STATUSES]
    missing = [finding for finding in findings if finding.status in MISSING_STATUSES]
    reviews = [finding for finding in findings if finding.status in REVIEW_STATUSES]
    if errors:
        conclusion = "发现异常"
        issues = errors + missing + reviews
    elif missing:
        conclusion = "存在缺填/未识别项"
        issues = missing + reviews
    elif reviews:
        conclusion = "待人工复核"
        issues = reviews
    else:
        conclusion = "比对通过"
        issues = []
    issue_text = "；".join(_short_finding(finding) for finding in issues[:10])
    if len(issues) > 10:
        issue_text += f"；另有{len(issues) - 10}项，详见内容核对报告"
    if not issue_text:
        issue_text = "申报表、身份证、学历及工作经历关联信息比对未发现异常"
    return {"结论": conclusion, "事项": issue_text}


def _finding_domain(finding: Finding) -> str:
    identity_fields = {"姓名", "性别", "身份证号", "出生日期", "身份证有效期至"}
    education_fields = {"最高学历", "学历层次", "学籍状态", "毕业院校", "毕业时间", "毕业证编码", "初中经历"}
    work_fields = {
        "工作经历", "工作经历解析", "申报表工作经历", "从事本职业年限", "累计工作年限",
        "工作起止时间", "工作经历重叠", "第一份工作", "最后一份工作", "首次工作年龄",
        "毕业后参加工作", "证明人姓名、电话", "证明人及电话", "工作年限", "全部工作月份",
        "企业全称", "企业信息截图", "从事职业", "单位公章", "规范工作陈述",
    }
    if finding.field in identity_fields or "身份" in finding.category:
        return "身份信息"
    if finding.field in education_fields or any(word in finding.category for word in ("学历", "学习经历")):
        return "学习经历"
    if finding.field in work_fields or any(word in finding.category for word in ("工作", "企业", "经营范围", "时间逻辑", "承诺函")):
        return "工作经历"
    return "其他审核"


def _domain_rows(result: PersonResult):
    grouped: dict[str, list[Finding]] = defaultdict(list)
    for finding in _content_findings(result):
        grouped[_finding_domain(finding)].append(finding)
    for domain in ("身份信息", "学习经历", "工作经历", "其他审核"):
        findings = grouped.get(domain, [])
        if not findings:
            yield [result.person, domain, "不适用", "本次未触发该类内容核对", "", ""]
            continue
        errors = [finding for finding in findings if finding.status in ERROR_STATUSES]
        missing = [finding for finding in findings if finding.status in MISSING_STATUSES]
        reviews = [finding for finding in findings if finding.status in REVIEW_STATUSES]
        passes = [finding for finding in findings if finding.status in PASS_STATUSES]
        if errors:
            conclusion = "存在异常"
        elif missing:
            conclusion = "存在缺填/未识别项"
        elif reviews:
            conclusion = "待人工复核"
        else:
            conclusion = "比对通过"
        facts = "；".join(_short_finding(finding) for finding in passes[:8])
        issues = "；".join(_short_finding(finding) for finding in (errors + missing + reviews)[:10])
        sources = "；".join(_unique([finding.sources for finding in findings if finding.sources])[:8])
        yield [result.person, domain, conclusion, facts or "已执行相关字段核对", issues, sources]


def _finish(ws) -> None:
    for row in ws.iter_rows(min_row=5):
        ws.row_dimensions[row[0].row].height = 34
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BOTTOM
            if str(cell.value) in STATUS_FILL:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL[str(cell.value)])
                cell.font = Font(name="Microsoft YaHei UI", size=10, color=TEXT, bold=True)
    for column in range(1, ws.max_column + 1):
        header = str(ws.cell(4, column).value or "")
        if header in WIDTHS:
            width = WIDTHS[header]
        else:
            values = [str(ws.cell(row, column).value or "") for row in range(4, min(ws.max_row, 80) + 1)]
            width = min(42, max(12, max((len(value) for value in values), default=12) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{max(4, ws.max_row)}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:4"


def write_report(results: list[PersonResult], output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    total = _sheet(
        wb,
        "人员核验总表",
        ["人员", "综合结论", "完整性结论", "缺少/无合格材料", "内容核对结论", "内容异常或复核事项", "采用材料数", "材料总数"],
        "一人一行查看最终结论。多份同类材料只要至少1份合格可用，完整性即按齐全处理；异常副本另列在材料清单。",
    )
    complete_summary = _sheet(
        wb,
        "完整性报告",
        ["人员", "完整性结论", "缺少/无合格材料", "采用材料数", "材料总数", "完整性说明"],
        "只判断需要提交的材料是否至少有1份合格可用；同类多份中的模糊、过期或读取失败副本不会推翻已存在的合格材料。",
    )
    content = _sheet(
        wb,
        "内容核对报告",
        ["人员", "核对项目", "结论", "核对事实", "异常/复核事项", "来源文件与页码"],
        "按身份信息、学习经历、工作经历和其他审核归并。无异常显示“比对通过”；错填、漏填、未识别和需人工复核事项如实列出。",
    )
    detail = _sheet(
        wb,
        "审核事实明细",
        ["人员", "类别", "字段", "状态", "说明", "字段值", "来源文件与页码"],
        "保留每一项规则的核验事实、标准化字段值和来源，便于追溯总表结论。",
    )
    review = _sheet(
        wb,
        "人工复核清单",
        ["人员", "类别", "字段", "状态", "复核原因", "识别值", "来源文件与页码"],
        "仅列出工具无法可靠自动判断的项目，人工确认后可结合审核事实明细作最终处理。",
    )
    evidence_sheet = _sheet(
        wb,
        "字段识别依据",
        ["人员", "材料类型", "字段", "识别值", "OCR置信度", "来源文件", "页码"],
        "展示工具实际识别到的字段。字段为空时请查看人工复核清单和材料清单中的处理异常。",
    )
    complete_detail = _sheet(
        wb,
        "材料完整性明细",
        ["人员", "材料类型", "状态", "判断依据", "数量或结果", "来源"],
        "逐项展示必需材料、条件材料和不适用材料的判断依据。",
    )
    works = _sheet(
        wb,
        "工作经历明细",
        ["人员", "企业名称", "从事职业", "开始时间", "结束时间", "工作月数", "证明人姓名", "证明人电话", "企业名称状态", "企业名称说明", "经营范围", "来源"],
        "展示从申报表读取并用于计算、与工作证明/承诺书/企业截图核对的结构化工作经历。",
    )
    mats = _sheet(
        wb,
        "材料清单",
        ["人员", "文件", "材料类型", "是否核验依据", "质量状态", "质量原因", "处理异常"],
        "列出全部文件。多份同类材料中未采用的副本仍保留记录，但不影响已存在合格材料的完整性结论。",
    )
    rejects = _sheet(
        wb,
        "异常与退回清单",
        ["人员", "类别", "字段/材料", "状态", "退回/异常原因", "字段值", "来源"],
        "集中列出不一致、缺少材料、退回和其他明确异常；待人工判断的事项另见人工复核清单。",
    )

    for result in results:
        material_summary = _material_summary(result)
        content_summary = _content_summary(result)
        summary = result.summary
        total.append([
            result.person,
            summary["总体结果"],
            material_summary["结论"],
            material_summary["缺少"],
            content_summary["结论"],
            content_summary["事项"],
            material_summary["采用数"],
            material_summary["总数"],
        ])
        complete_summary.append([
            result.person,
            material_summary["结论"],
            material_summary["缺少"],
            material_summary["采用数"],
            material_summary["总数"],
            material_summary["说明"],
        ])
        for row in _domain_rows(result):
            content.append(row)
        for finding in result.findings:
            detail.append([
                finding.person, finding.category, finding.field, finding.status,
                finding.message, finding.values, finding.sources,
            ])
            if finding.status in REVIEW_STATUSES | MISSING_STATUSES:
                review.append([
                    finding.person, finding.category, finding.field, finding.status,
                    finding.message, finding.values, finding.sources,
                ])
            if finding.category == "材料完整性":
                complete_detail.append([
                    finding.person, finding.field, finding.status,
                    finding.message, finding.values, finding.sources,
                ])
            if finding.status in ERROR_STATUSES:
                rejects.append([
                    finding.person, finding.category, finding.field, finding.status,
                    finding.message, finding.values, finding.sources,
                ])
        for work in result.work_records:
            works.append([
                work.person, work.company, work.occupation, work.start, work.end,
                work.duration_months, work.witness_name, work.witness_phone,
                work.company_status, work.company_message, work.business_scope, work.source,
            ])
        for material in result.materials:
            mats.append([
                material.person, material.path.name, material.document_type,
                "是" if material.selected_as_basis else "否", material.quality_status,
                "；".join(material.quality_reasons), "；".join(material.errors),
            ])
            for evidence in material.evidences:
                confidence = "" if evidence.confidence is None else evidence.confidence
                evidence_sheet.append([
                    evidence.person, evidence.document_type, evidence.field,
                    evidence.normalized_value, confidence, evidence.file, evidence.page,
                ])

    evidence_sheet.column_dimensions["E"].width = 13
    for row in range(5, evidence_sheet.max_row + 1):
        if isinstance(evidence_sheet.cell(row, 5).value, float):
            evidence_sheet.cell(row, 5).number_format = "0%"
    for ws in wb.worksheets:
        _finish(ws)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
