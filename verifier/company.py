from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from .normalize import normalize_company

ORG_SUFFIXES = (
    "有限责任公司", "股份有限公司", "有限公司", "集团有限公司",
    "合伙企业", "个人独资企业", "分公司", "支公司", "事务所",
    "研究院", "事业单位", "社会团体",
)


class CompanyRegistry:
    def __init__(self, path: str | None = None):
        self.names: set[str] = set()
        if path:
            self.load(Path(path))

    def load(self, path: Path) -> None:
        values: list[str] = []
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
                values = [r[0] for r in rows[1:] if r]
        else:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            values = [str(r[0].value) for r in ws.iter_rows(min_row=2) if r[0].value]
        self.names = {normalize_company(x) for x in values if normalize_company(x)}

    @staticmethod
    def looks_full(name: str) -> bool:
        n = normalize_company(name)
        return len(n) >= 6 and any(n.endswith(s) for s in ORG_SUFFIXES)

    def validate(self, name: str) -> tuple[str, str]:
        n = normalize_company(name)
        if not self.looks_full(n):
            return "退回", "该企业信息为简称，请补充与工商信息一致的全称"
        if self.names:
            if n in self.names:
                return "通过", "与本地工商全称名录完全一致"
            return "待复核", "无法确认是否为工商登记全称，请补充证明或更新企业名录"
        return "待复核", "名称形式完整，但未提供本地工商全称名录，无法确认登记一致性"

